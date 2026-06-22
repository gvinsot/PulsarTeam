"""Excel (.xlsx) — read, edit-in-place, generate. Backed by openpyxl.

Editing is round-trip (load workbook, set the targeted cells, save) so styles,
other sheets, and formulas the LLM didn't touch are preserved. The LLM emits
small `{sheet, cell, value}` deltas. Reads render each sheet as a markdown table
plus a structured cell grid.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from ..util import ToolError, check_size, default_output, resolve_output, resolve_path

# Cap read width/height so a huge sheet can't blow up the LLM context.
MAX_ROWS = 500
MAX_COLS = 50


def _sheet_to_markdown(ws, max_rows: int, max_cols: int) -> str:
    rows = []
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if r >= max_rows:
            rows.append(f"... ({ws.max_row - max_rows} more rows)")
            break
        cells = ["" if v is None else str(v) for v in row[:max_cols]]
        rows.append("| " + " | ".join(cells) + " |")
    if not rows:
        return "_(empty sheet)_"
    # Insert a markdown header separator after the first row.
    if rows and rows[0].startswith("|"):
        ncols = rows[0].count("|") - 1
        rows.insert(1, "|" + "|".join([" --- "] * ncols) + "|")
    return "\n".join(rows)


def read_xlsx(path: str, sheet: str | None = None, max_rows: int = MAX_ROWS, max_cols: int = MAX_COLS) -> dict[str, Any]:
    """Return sheet dimensions, a cell grid, and markdown per sheet."""
    src = resolve_path(path)
    check_size(src)
    wb = load_workbook(str(src), data_only=False)
    names = wb.sheetnames
    targets = [sheet] if sheet else names
    sheets = []
    for name in targets:
        if name not in wb:
            raise ToolError(f"sheet not found: {name} (available: {names})")
        ws = wb[name]
        grid = []
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r >= max_rows:
                break
            grid.append(["" if v is None else v for v in row[:max_cols]])
        sheets.append({
            "name": name,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "grid": grid,
            "markdown": _sheet_to_markdown(ws, max_rows, max_cols),
        })
    return {"format": "xlsx", "path": str(src), "sheet_names": names, "sheets": sheets}


def get_outline(path: str) -> dict[str, Any]:
    src = resolve_path(path)
    wb = load_workbook(str(src), read_only=True)
    sheets = [{"name": ws.title, "rows": ws.max_row, "cols": ws.max_column} for ws in wb.worksheets]
    return {"format": "xlsx", "path": str(src), "sheets": sheets}


def edit_xlsx(path: str, cells: list[dict[str, Any]], output_path: str | None = None) -> dict[str, Any]:
    """Set cells. Each item: {sheet?, cell:"B3", value} ; value starting with '='
    is written as a formula. Default sheet = active sheet."""
    src = resolve_path(path)
    if not isinstance(cells, list) or not cells:
        raise ToolError("cells must be a non-empty list of {sheet?, cell, value}")
    wb = load_workbook(str(src))
    applied = 0
    for item in cells:
        name = item.get("sheet")
        ws = wb[name] if name else wb.active
        if name and name not in wb:
            raise ToolError(f"sheet not found: {name}")
        ref = item.get("cell")
        if not ref:
            raise ToolError("each cell item needs 'cell' (e.g. 'B3')")
        ws[ref] = item.get("value")
        applied += 1
    out = resolve_output(output_path) if output_path else default_output(src)
    wb.save(str(out))
    return {"ok": True, "applied": applied, "output_path": str(out)}


def generate_xlsx(output_path: str, spec: dict[str, Any]) -> dict[str, Any]:
    """Create a new workbook.

    spec = {"sheets": [{"name": "Sheet1", "header": [...]?, "rows": [[...], ...]}]}
    A single-sheet shortcut {"header": [...], "rows": [...]} is also accepted.
    """
    out = resolve_output(output_path)
    wb = Workbook()
    sheets = spec.get("sheets")
    if sheets is None and ("rows" in spec or "header" in spec):
        sheets = [{"name": "Sheet1", "header": spec.get("header"), "rows": spec.get("rows", [])}]
    if not sheets:
        raise ToolError("spec needs 'sheets' or a top-level 'rows'/'header'")
    first = True
    for s in sheets:
        ws = wb.active if first else wb.create_sheet()
        if s.get("name"):
            ws.title = str(s["name"])[:31]  # Excel sheet-name limit
        first = False
        if s.get("header"):
            ws.append([str(h) for h in s["header"]])
        for row in s.get("rows", []) or []:
            ws.append(list(row))
    wb.save(str(out))
    return {"ok": True, "output_path": str(out)}
